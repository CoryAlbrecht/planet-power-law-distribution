"""Format and export the Excel workbook."""

from typing import cast

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from planet_power.constants import (
    COLUMN_GROUPS,
    G_EARTH,
    GROUP_COLOURS,
    M_JUP_KG,
    R_JUP_M,
    G,
)


def format_workbook(path: str, n_rows: int) -> None:
    """Apply professional formatting to the saved workbook."""
    wb = load_workbook(path)
    ws = cast(Worksheet, wb.active)
    ws.title = "Exoplanets"

    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    data_font = Font(name="Arial", size=9)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align_c = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    ws.row_dimensions[1].height = 48
    for col_idx, cell in enumerate(ws[1], start=1):
        col_name: str = cell.value or ""  # type: ignore[reportUnknownVariableType]
        group: str = COLUMN_GROUPS.get(col_name, "Unknown")
        grp_hex: str = GROUP_COLOURS.get(group, "E0E0E0")
        r = int(grp_hex[0:2], 16)
        g = int(grp_hex[2:4], 16)
        b = int(grp_hex[4:6], 16)
        dark_hex = f"{max(r-60,0):02X}{max(g-60,0):02X}{max(b-60,0):02X}"
        cell.font = header_font
        cell.fill = PatternFill("solid", start_color=dark_hex)
        cell.alignment = header_align
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
        for cell in row:
            col_name: str = ws.cell(row=1, column=cell.column).value or ""  # type: ignore[reportUnknownMemberType]
            group: str = COLUMN_GROUPS.get(col_name, "Unknown")  # type: ignore[reportUnknownArgumentType]
            grp_hex: str = GROUP_COLOURS.get(group, "FFFFFF")
            cell.font = data_font
            cell.alignment = data_align_c
            cell.border = thin_border
            if cell.row % 2 == 0:  # type: ignore[reportOptionalOperand]
                cell.fill = PatternFill("solid", start_color=grp_hex)

        ws.row_dimensions[cell.row].height = 15  # type: ignore[reportPossiblyUnboundVariable]

    col_widths = {
        "Planet Name": 22,
        "Host Star": 18,
        "Letter": 8,
        "Discovery Method": 22,
        "Discovery Year": 12,
        "Discovery Facility": 24,
        "Orbital Period (days)": 16,
        "Semi-Major Axis (AU)": 16,
    }
    num_fmt_cols = {
        "Mass (kg)",
        "Mass Err+ (kg)",
        "Mass Err- (kg)",
        "Radius (m)",
        "Radius Err+ (m)",
        "Radius Err- (m)",
    }
    for col_idx, cell in enumerate(ws[1], start=1):
        col_letter = get_column_letter(col_idx)
        name: str = cell.value or ""  # type: ignore[reportUnknownMemberType]
        if name in col_widths:
            ws.column_dimensions[col_letter].width = col_widths[name]  # type: ignore[reportArgumentType]
        elif name in num_fmt_cols:
            ws.column_dimensions[col_letter].width = 18
        elif "Err" in name or "Unc" in name:  # type: ignore[reportOperatorIssue]
            ws.column_dimensions[col_letter].width = 14
        else:
            ws.column_dimensions[col_letter].width = 16

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = ws.dimensions

    notes = wb.create_sheet("Notes")
    notes_content: list[list[str]] = [
        ["Exoplanet Dataset — Notes"],
        [],
        ["Source", "NASA Exoplanet Archive — PSCompPars table"],
        ["Table", "Planetary Systems Composite Parameters (pscomppars)"],
        [
            "Filter",
            "pl_bmassj IS NOT NULL AND pl_radj IS NOT NULL AND pl_dens IS NOT NULL",
        ],
        ["URL", "https://exoplanetarchive.ipac.caltech.edu/TAP/sync"],
        [],
        ["Computed columns"],
        ["Mass (kg)", "pl_bmassj * M_Jup (kg)"],
        ["Mass Err+ (kg)", "pl_bmassjerr1 * M_Jup (kg)"],
        ["Mass Err- (kg)", "pl_bmassjerr2 * M_Jup (kg)"],
        ["Radius (m)", "pl_radj * R_Jup (m)"],
        ["Radius Err+ (m)", "pl_radjerr1 * R_Jup (m)"],
        ["Radius Err- (m)", "pl_radjerr2 * R_Jup (m)"],
        ["Surface Gravity (m/s²)", "g = G * (pl_bmassj * M_Jup) / (pl_radj * R_Jup)^2"],
        ["Surface Gravity (g_Earth)", "g / 9.80665"],
        ["Surf. Grav. Err+ (m/s²)", "g(M+dmass, R+drad) - g(M,R)"],
        ["Surf. Grav. Err- (m/s²)", "g(M,R) - g(M-dmass, R-drad)"],
        ["Surf. Grav. Err+ (g_Earth)", "Err+ (m/s²) / 9.80665"],
        ["Surf. Grav. Err- (g_Earth)", "Err- (m/s²) / 9.80665"],
        [],
        ["Durand-Manterola (2011) Classification"],
        [
            "Reference",
            "Durand-Manterola, H.J. (2011). Planets: Power Laws and Classification. arXiv:1111.3986",
        ],
        [
            "DM Class A",
            "M < 5×10²⁵ kg  (~0.026 M_Jup, ~8.3 M_Earth) — rocky / terrestrial",
        ],
        [
            "DM Class B",
            "5×10²⁵ ≤ M < 10²⁷ kg  (~0.53 M_Jup, ~167 M_Earth) — volatile-rich / ice giants",
        ],
        ["DM Class C", "M ≥ 10²⁷ kg  (>0.53 M_Jup) — Jovian gas giants"],
        ["DM Predicted g", "Class A: g = 2×10⁻¹⁰ × M^0.4282  (eq. 4a)"],
        ["", "Class B: g = 14937 × M^(-0.1219)  (eq. 4b)"],
        ["", "Class C: g = 4×10⁻²⁸ × M^1.0482  (eq. 4c)"],
        ["", "M in kg, g in m s⁻²"],
        [
            "DM Grav. Residual",
            "Computed g_Earth minus DM-predicted g_Earth; large values flag unusual planets",
        ],
        [],
        ["Constants used"],
        ["G", f"{G} m³ kg⁻¹ s⁻²"],
        ["M_Jupiter", f"{M_JUP_KG} kg"],
        ["R_Jupiter", f"{R_JUP_M} m"],
        ["g_Earth", f"{G_EARTH} m s⁻²"],
        [],
        ["Notes on PSCompPars"],
        ["", "One row per confirmed planet."],
        [
            "",
            "Parameters may come from different references (not always self-consistent).",
        ],
        [
            "",
            "Density values marked 'calculated' were derived by NASA from mass+radius.",
        ],
        [
            "",
            "Filter out calculated values by checking the pl_dens_reflink column (not included here).",
        ],
        [],
        ["Citation"],
        [
            "",
            "NASA Exoplanet Archive. (2024). Planetary Systems Composite Parameters Table.",
        ],
        ["", "DOI: 10.26133/NEA12"],
    ]
    for row in notes_content:  # type: ignore[reportUnknownVariableType]
        notes.append(row)
    notes.column_dimensions["A"].width = 26
    notes.column_dimensions["B"].width = 72
    for cell in notes["A"]:
        cell.font = Font(name="Arial", size=10, bold=(cell.row == 1))
    for cell in notes["B"]:
        cell.font = Font(name="Arial", size=10)
    notes["A1"].font = Font(name="Arial", size=13, bold=True)

    wb.save(path)
